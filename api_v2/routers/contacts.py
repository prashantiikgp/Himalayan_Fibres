"""Contacts endpoints — Phase 1 list/filter/segments.

This commit ships the read-side: GET /contacts (paginated, filterable),
GET /contacts/tags, GET /contacts/countries, GET /segments.

Edit + import + drawer detail land in a follow-up commit.

Implementation reuses hf_dashboard/services/segments.py helpers — no
duplicated business logic. The "segments matched" column is computed
in Python for the current page only (same trick v1 uses; ~0ms for
50 rows × ~10 segments).
"""

from __future__ import annotations

from typing import Annotated, Literal

import csv
import io
import uuid
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import or_

from api_v2.deps import require_auth
from api_v2.schemas.contacts import (
    ContactCreate,
    ContactDetail,
    ContactInteractionOut,
    ContactListResponse,
    ContactNoteOut,
    ContactRow,
    ContactUpdate,
    CountriesResponse,
    ImportResponse,
    NoteCreate,
    SegmentSummary,
    SegmentsResponse,
    TagsResponse,
)

# Reused v1 services — single source for segment rules + caching.
from services.database import get_db  # type: ignore[import-not-found]
from services.interactions import (  # type: ignore[import-not-found]
    get_interactions,
    log_interaction,
    summarize_diff,
)
from services.models import (  # type: ignore[import-not-found]
    Contact,
    ContactNote,
)
from services.segments import (  # type: ignore[import-not-found]
    count_segment_members,
    get_active_segments_cached,
    get_all_tags_from_contacts,
    segments_for_contact,
)


def _wa_id_from_phone(phone: str) -> str | None:
    """v1 logic: 10-digit Indian → 91XXX..., longer → strip +."""
    digits = "".join(ch for ch in phone if ch.isdigit())
    if len(digits) == 10:
        return f"91{digits}"
    if len(digits) > 10:
        return digits.lstrip("+")
    return None

router = APIRouter()


def _is_real_email(email: str | None) -> bool:
    if not email:
        return False
    if "@placeholder.local" in email or email.startswith("wa_"):
        return False
    return True


@router.get("/contacts", response_model=ContactListResponse)
async def list_contacts(
    _auth: Annotated[None, Depends(require_auth)],
    segment: str | None = Query(None, description="Filter by customer_type segment id"),
    lifecycle: str | None = Query(None, description="Filter by lifecycle stage id"),
    country: str | None = Query(None),
    channel: Literal["all", "email", "whatsapp", "both"] = "all",
    tags: list[str] = Query(default_factory=list),
    search: str = "",
    page: int = Query(0, ge=0),
    page_size: int = Query(50, ge=1, le=200),
) -> ContactListResponse:
    db = get_db()
    try:
        # Plan D Phase 1.3 — select only the 14 columns the row renderer +
        # rule engine touch, not the full ~38 column Contact row.
        q = db.query(Contact).with_entities(
            Contact.id,
            Contact.first_name,
            Contact.last_name,
            Contact.company,
            Contact.email,
            Contact.phone,
            Contact.wa_id,
            Contact.lifecycle,
            Contact.customer_type,
            Contact.consent_status,
            Contact.country,
            Contact.tags,
            Contact.customer_subtype,
            Contact.geography,
        )

        if segment and segment != "all":
            q = q.filter(Contact.customer_type == segment)
        if lifecycle and lifecycle != "all":
            q = q.filter(Contact.lifecycle == lifecycle)
        if country and country != "all":
            q = q.filter(Contact.country == country)
        if channel == "email":
            q = q.filter(Contact.email.isnot(None), ~Contact.email.like("%placeholder%"))
        elif channel == "whatsapp":
            q = q.filter(Contact.wa_id.isnot(None))
        elif channel == "both":
            q = q.filter(
                Contact.email.isnot(None),
                ~Contact.email.like("%placeholder%"),
                Contact.wa_id.isnot(None),
            )
        if search:
            term = f"%{search}%"
            q = q.filter(
                or_(
                    Contact.email.ilike(term),
                    Contact.first_name.ilike(term),
                    Contact.last_name.ilike(term),
                    Contact.company.ilike(term),
                )
            )

        # Tag filter — JSON columns evaluated in Python (SQLite + Postgres compat).
        tag_set = {t.strip() for t in tags if t and t.strip()} or None
        if tag_set:
            id_subset: set[str] = set()
            for cid, ctags in db.query(Contact.id, Contact.tags).all():
                ctags_list = ctags or []
                if any(t in tag_set for t in ctags_list):
                    id_subset.add(cid)
            if not id_subset:
                return ContactListResponse(
                    contacts=[], total=0, page=0, page_size=page_size, total_pages=1
                )
            q = q.filter(Contact.id.in_(id_subset))

        total = q.count()
        total_pages = max(1, (total + page_size - 1) // page_size)
        if page >= total_pages:
            page = max(total_pages - 1, 0)

        rows = q.order_by(Contact.company, Contact.first_name).offset(page * page_size).limit(page_size).all()

        # Segment-rule eval for the page only — cached segments list (5min TTL).
        all_segments = get_active_segments_cached()
        contacts_out: list[ContactRow] = []
        for r in rows:
            seg_ids = segments_for_contact(r, all_segments)
            channels: list[Literal["email", "whatsapp"]] = []
            if _is_real_email(r.email):
                channels.append("email")
            if r.wa_id:
                channels.append("whatsapp")
            contacts_out.append(
                ContactRow(
                    id=r.id,
                    first_name=r.first_name or "",
                    last_name=r.last_name or "",
                    company=r.company or "",
                    email=r.email if _is_real_email(r.email) else "",
                    phone=r.phone or "",
                    wa_id=r.wa_id,
                    lifecycle=r.lifecycle or "",
                    customer_type=r.customer_type or "",
                    consent_status=r.consent_status or "",
                    country=r.country or "",
                    tags=list(r.tags or []),
                    segments=seg_ids,
                    channels=channels,
                )
            )

        return ContactListResponse(
            contacts=contacts_out,
            total=total,
            page=page,
            page_size=page_size,
            total_pages=total_pages,
        )
    finally:
        db.close()


def _iso(dt) -> str:  # type: ignore[no-untyped-def]
    if dt is None:
        return ""
    if dt.tzinfo is None:
        from datetime import timezone

        dt = dt.replace(tzinfo=timezone.utc)
    return dt.isoformat()


# NOTE on route order: FastAPI dispatches in registration order. The static
# `/contacts/tags` and `/contacts/countries` routes MUST be declared before
# the dynamic `/contacts/{contact_id}` — otherwise "tags" gets matched as
# a contact_id and 404s on the lookup.
@router.get("/contacts/tags", response_model=TagsResponse)
async def list_contact_tags(_auth: Annotated[None, Depends(require_auth)]) -> TagsResponse:
    db = get_db()
    try:
        return TagsResponse(tags=get_all_tags_from_contacts(db))
    finally:
        db.close()


@router.get("/contacts/countries", response_model=CountriesResponse)
async def list_contact_countries(
    _auth: Annotated[None, Depends(require_auth)],
) -> CountriesResponse:
    db = get_db()
    try:
        rows = (
            db.query(Contact.country)
            .filter(Contact.country.isnot(None), Contact.country != "")
            .distinct()
            .order_by(Contact.country)
            .all()
        )
        return CountriesResponse(countries=[r[0] for r in rows if r[0]])
    finally:
        db.close()


@router.get("/contacts/{contact_id}", response_model=ContactDetail)
async def get_contact(
    contact_id: str,
    _auth: Annotated[None, Depends(require_auth)],
) -> ContactDetail:
    """Full contact detail — drives the ContactDrawer on the Contacts page."""
    db = get_db()
    try:
        c = db.query(Contact).filter(Contact.id == contact_id).one_or_none()
        if c is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Contact not found")

        all_segments = get_active_segments_cached()
        seg_ids = segments_for_contact(c, all_segments)
        by_id = {s.id: s for s in all_segments}
        matched = [
            SegmentSummary(
                id=s.id,
                name=s.name,
                color=getattr(s, "color", None),
                description=getattr(s, "description", None),
                member_count=count_segment_members(db, s),
            )
            for sid in seg_ids
            if (s := by_id.get(sid)) is not None
        ]

        threaded_rows = (
            db.query(ContactNote)
            .filter(ContactNote.contact_id == c.id)
            .order_by(ContactNote.created_at.desc())
            .all()
        )
        threaded_notes = [
            ContactNoteOut(id=n.id, body=n.body, author=n.author, created_at=_iso(n.created_at))
            for n in threaded_rows
        ]

        try:
            activity_rows = get_interactions(db, c.id, limit=50)
        except Exception:
            activity_rows = []
        activity = [
            ContactInteractionOut(
                id=a.id,
                kind=a.kind,
                summary=a.summary or "",
                actor=getattr(a, "actor", None),
                created_at=_iso(a.created_at),
            )
            for a in activity_rows
        ]

        channels: list[Literal["email", "whatsapp"]] = []
        if _is_real_email(c.email):
            channels.append("email")
        if c.wa_id:
            channels.append("whatsapp")

        return ContactDetail(
            id=c.id,
            first_name=c.first_name or "",
            last_name=c.last_name or "",
            company=c.company or "",
            email=c.email if _is_real_email(c.email) else "",
            phone=c.phone or "",
            wa_id=c.wa_id,
            lifecycle=c.lifecycle or "",
            customer_type=c.customer_type or "",
            consent_status=c.consent_status or "",
            country=c.country or "",
            tags=list(c.tags or []),
            segments=seg_ids,
            channels=channels,
            customer_subtype=c.customer_subtype or "",
            geography=c.geography or "",
            legacy_notes=c.notes or "",
            threaded_notes=threaded_notes,
            activity=activity,
            matched_segments=matched,
        )
    finally:
        db.close()


@router.post("/contacts", response_model=ContactRow, status_code=status.HTTP_201_CREATED)
async def create_contact(
    body: ContactCreate,
    _auth: Annotated[None, Depends(require_auth)],
) -> ContactRow:
    """Create a new contact. Required: first_name + phone.

    Mirrors v1's Add Contact behavior — generates an 8-char UUID id, derives
    wa_id from the phone, defaults consent to 'pending'.
    """
    db = get_db()
    try:
        clean_phone = "".join(ch for ch in body.phone if ch.isdigit())
        if len(clean_phone) < 10:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Phone must contain at least 10 digits",
            )
        wa_id = _wa_id_from_phone(body.phone)

        if body.email:
            existing = db.query(Contact).filter(Contact.email == body.email).first()
            if existing:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=f"Email already exists: {body.email}",
                )

        new_id = str(uuid.uuid4())[:8]
        contact = Contact(
            id=new_id,
            first_name=body.first_name,
            last_name=body.last_name,
            phone=clean_phone,
            email=body.email or f"wa_{wa_id}@placeholder.local",
            company=body.company,
            country=body.country,
            customer_type=body.customer_type,
            lifecycle=body.lifecycle,
            tags=body.tags,
            wa_id=wa_id,
            consent_status="pending",
        )
        db.add(contact)
        db.commit()

        try:
            log_interaction(
                db,
                contact_id=new_id,
                kind="imported",
                summary=f"Added via API · {body.first_name} {body.last_name}".strip(),
                actor="api_v2",
            )
        except Exception:
            pass  # interaction log is best-effort

        all_segments = get_active_segments_cached()
        seg_ids = segments_for_contact(contact, all_segments)
        channels: list[Literal["email", "whatsapp"]] = []
        if _is_real_email(contact.email):
            channels.append("email")
        if contact.wa_id:
            channels.append("whatsapp")

        return ContactRow(
            id=contact.id,
            first_name=contact.first_name,
            last_name=contact.last_name,
            company=contact.company,
            email=contact.email if _is_real_email(contact.email) else "",
            phone=contact.phone,
            wa_id=contact.wa_id,
            lifecycle=contact.lifecycle,
            customer_type=contact.customer_type,
            consent_status=contact.consent_status,
            country=contact.country,
            tags=list(contact.tags or []),
            segments=seg_ids,
            channels=channels,
        )
    finally:
        db.close()


@router.patch("/contacts/{contact_id}", response_model=ContactRow)
async def update_contact(
    contact_id: str,
    body: ContactUpdate,
    _auth: Annotated[None, Depends(require_auth)],
) -> ContactRow:
    """Edit a contact. All fields optional; only provided fields are updated.

    Logs a `manual_edit` interaction with a human-readable diff summary so
    the Activity tab in the drawer reflects the change.
    """
    db = get_db()
    try:
        c = db.query(Contact).filter(Contact.id == contact_id).one_or_none()
        if c is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Contact not found")

        # Email uniqueness check (only if changing to a new value)
        if body.email is not None and body.email and body.email != c.email:
            other = (
                db.query(Contact)
                .filter(Contact.email == body.email, Contact.id != contact_id)
                .first()
            )
            if other:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=f"Email already in use: {body.email}",
                )

        before: dict[str, Any] = {
            "first_name": c.first_name,
            "last_name": c.last_name,
            "phone": c.phone,
            "email": c.email,
            "company": c.company,
            "country": c.country,
            "lifecycle": c.lifecycle,
            "consent_status": c.consent_status,
            "tags": list(c.tags or []),
            "notes": c.notes,
        }

        if body.first_name is not None:
            c.first_name = body.first_name.strip()
        if body.last_name is not None:
            c.last_name = body.last_name.strip()
        if body.phone is not None:
            c.phone = "".join(ch for ch in body.phone if ch.isdigit())
            wa_id = _wa_id_from_phone(c.phone)
            if wa_id:
                c.wa_id = wa_id
        if body.email is not None:
            c.email = body.email.strip() if body.email.strip() else c.email
        if body.company is not None:
            c.company = body.company.strip()
        if body.country is not None:
            c.country = body.country.strip() or "India"
        if body.lifecycle is not None:
            c.lifecycle = body.lifecycle
        if body.consent_status is not None:
            c.consent_status = body.consent_status
        if body.tags is not None:
            c.tags = [t.strip() for t in body.tags if t and t.strip()]
        if body.notes is not None:
            c.notes = body.notes

        db.commit()

        after = {
            "first_name": c.first_name,
            "last_name": c.last_name,
            "phone": c.phone,
            "email": c.email,
            "company": c.company,
            "country": c.country,
            "lifecycle": c.lifecycle,
            "consent_status": c.consent_status,
            "tags": list(c.tags or []),
            "notes": c.notes,
        }
        try:
            diff_summary = summarize_diff(before, after)
            if diff_summary != "no-op save":
                log_interaction(
                    db,
                    contact_id=contact_id,
                    kind="manual_edit",
                    summary=f"Changed: {diff_summary}",
                    actor="api_v2",
                )
        except Exception:
            pass

        all_segments = get_active_segments_cached()
        seg_ids = segments_for_contact(c, all_segments)
        channels: list[Literal["email", "whatsapp"]] = []
        if _is_real_email(c.email):
            channels.append("email")
        if c.wa_id:
            channels.append("whatsapp")

        return ContactRow(
            id=c.id,
            first_name=c.first_name or "",
            last_name=c.last_name or "",
            company=c.company or "",
            email=c.email if _is_real_email(c.email) else "",
            phone=c.phone or "",
            wa_id=c.wa_id,
            lifecycle=c.lifecycle or "",
            customer_type=c.customer_type or "",
            consent_status=c.consent_status or "",
            country=c.country or "",
            tags=list(c.tags or []),
            segments=seg_ids,
            channels=channels,
        )
    finally:
        db.close()


@router.post("/contacts/{contact_id}/notes", response_model=ContactNoteOut, status_code=201)
async def add_contact_note(
    contact_id: str,
    body: NoteCreate,
    _auth: Annotated[None, Depends(require_auth)],
) -> ContactNoteOut:
    """Append a threaded note to a contact. Logs a `note_added` interaction."""
    db = get_db()
    try:
        c = db.query(Contact).filter(Contact.id == contact_id).one_or_none()
        if c is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Contact not found")

        note = ContactNote(contact_id=contact_id, body=body.body.strip(), author="user")
        db.add(note)
        db.commit()
        db.refresh(note)

        try:
            log_interaction(
                db,
                contact_id=contact_id,
                kind="note_added",
                summary=body.body[:120],
                actor="user",
            )
        except Exception:
            pass

        return ContactNoteOut(
            id=note.id, body=note.body, author=note.author, created_at=_iso(note.created_at)
        )
    finally:
        db.close()


@router.post("/contacts/import", response_model=ImportResponse)
async def import_contacts(
    _auth: Annotated[None, Depends(require_auth)],
    file: UploadFile = File(...),
) -> ImportResponse:
    """Bulk-import contacts from CSV or Excel.

    Expected columns (any of these — case-insensitive): email (required),
    first_name/name, last_name, company, phone, country.

    Skips rows with missing email or duplicate email. Returns counts + errors.
    """
    if not file.filename or not file.filename.lower().endswith((".csv", ".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Expected .csv, .xlsx, or .xls file")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    rows: list[dict[str, str]] = []
    errors: list[str] = []
    try:
        if file.filename.lower().endswith(".csv"):
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            for r in reader:
                rows.append({(k or "").lower().strip(): (v or "").strip() for k, v in r.items()})
        else:
            try:
                import openpyxl  # type: ignore[import-not-found]

                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                ws = wb.active
                if ws is None:
                    raise HTTPException(status_code=400, detail="Empty workbook")
                headers_row = next(ws.iter_rows(values_only=True), None)
                if not headers_row:
                    raise HTTPException(status_code=400, detail="No header row")
                headers = [str(h or "").lower().strip() for h in headers_row]
                for raw in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(
                        {
                            headers[i]: str(raw[i] or "").strip()
                            for i in range(min(len(headers), len(raw)))
                        }
                    )
            except ImportError:
                raise HTTPException(
                    status_code=500,
                    detail="openpyxl not installed; CSV imports work, Excel does not",
                )
    except UnicodeDecodeError as e:
        raise HTTPException(status_code=400, detail=f"Could not decode file as UTF-8: {e}")

    db = get_db()
    imported = 0
    skipped = 0
    try:
        for i, row in enumerate(rows, start=2):  # row 2 = first data row in CSV
            email = (row.get("email") or row.get("e-mail") or "").strip()
            if not email or "@" not in email:
                skipped += 1
                continue
            if db.query(Contact).filter(Contact.email == email).first():
                skipped += 1
                continue

            phone = (row.get("phone") or row.get("mobile") or "").strip()
            wa_id = _wa_id_from_phone(phone) if phone else None
            try:
                db.add(
                    Contact(
                        id=str(uuid.uuid4())[:8],
                        email=email,
                        first_name=row.get("first_name") or row.get("name") or "",
                        last_name=row.get("last_name") or "",
                        company=row.get("company") or "",
                        phone="".join(ch for ch in phone if ch.isdigit()),
                        country=row.get("country") or "India",
                        wa_id=wa_id,
                        consent_status="pending",
                        lifecycle="new_lead",
                    )
                )
                imported += 1
            except Exception as e:
                errors.append(f"Row {i}: {e}")
                skipped += 1
        db.commit()
    finally:
        db.close()

    return ImportResponse(imported=imported, skipped=skipped, errors=errors)


@router.get("/contacts.csv")
async def download_contacts_csv(
    _auth: Annotated[None, Depends(require_auth)],
) -> StreamingResponse:
    """Stream all contacts as CSV.

    Plan D Phase 1.2: pulls only the 9 columns the CSV uses, not the
    full 38-col Contact row.
    """

    def generate() -> Any:
        db = get_db()
        try:
            buf = io.StringIO()
            writer = csv.writer(buf)
            header = [
                "email",
                "first_name",
                "last_name",
                "company",
                "phone",
                "country",
                "lifecycle",
                "consent_status",
                "wa_id",
            ]
            writer.writerow(header)
            yield buf.getvalue()
            buf.seek(0)
            buf.truncate(0)

            rows = db.query(
                Contact.email,
                Contact.first_name,
                Contact.last_name,
                Contact.company,
                Contact.phone,
                Contact.country,
                Contact.lifecycle,
                Contact.consent_status,
                Contact.wa_id,
            ).yield_per(500)
            for r in rows:
                writer.writerow(
                    [
                        r.email or "",
                        r.first_name or "",
                        r.last_name or "",
                        r.company or "",
                        r.phone or "",
                        r.country or "",
                        r.lifecycle or "",
                        r.consent_status or "",
                        r.wa_id or "",
                    ]
                )
                yield buf.getvalue()
                buf.seek(0)
                buf.truncate(0)
        finally:
            db.close()

    return StreamingResponse(
        generate(),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="contacts.csv"'},
    )


@router.get("/segments", response_model=SegmentsResponse)
async def list_segments(_auth: Annotated[None, Depends(require_auth)]) -> SegmentsResponse:
    db = get_db()
    try:
        segments = get_active_segments_cached()
        out = [
            SegmentSummary(
                id=s.id,
                name=s.name,
                color=getattr(s, "color", None),
                description=getattr(s, "description", None),
                member_count=count_segment_members(db, s),
            )
            for s in segments
        ]
        return SegmentsResponse(segments=out)
    finally:
        db.close()
